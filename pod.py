import sublist3r
import csv
from openpyxl import load_workbook

def find_subdomains(domain):
    subdomains = sublist3r.main(domain, 40, None, ports=None, silent=True, verbose=False, enable_bruteforce=False, engines=None)
    return subdomains

def save_to_csv(domain, result, file):
    with open(file, 'a', newline='') as csvfile:
        writer = csv.writer(csvfile)
        writer.writerow([f"Subdomains for domain {domain}:"])
        for subdomain in result:
            writer.writerow([subdomain])

xlsx_file_name = "domains.xlsx"

# Load domains from XLS file
wb = load_workbook(xlsx_file_name)
sheet = wb.active
domains = []
for row in sheet.iter_rows(values_only=True):
    for cell in row:
        if cell is not None:
            domains.append(str(cell))

# Find subdomains and save results for each domain
for domain in domains:
    result = find_subdomains(domain)
    print("Found subdomains for {}: ".format(domain))
    for subdomain in result:
        print(subdomain)
    csv_file_name = f"{domain}_subdomains.csv"
    save_to_csv(domain, result, csv_file_name)
    print("Results for {} saved to file: {}".format(domain, csv_file_name))


